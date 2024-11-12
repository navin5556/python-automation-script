# To get all rows where a specific column matches a given value
import openpyxl 
wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet3']


column_letter = 'B'  # Change this to the column letter you want to search
search_value = 'Apples'  # Change this to the value you are looking for

for row in range(1, sheet.max_row + 1):
    if sheet[column_letter + str(row)].value == search_value:
        for col in range(1, sheet.max_column + 1):
            print(sheet.cell(row=row, column=col).value, end=" ")
        print()
