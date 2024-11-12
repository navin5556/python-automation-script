# To print a complete row, you can iterate through all the columns in that row

import openpyxl 
wb = openpyxl.load_workbook('example.xlsx')
# print(wb.sheetnames)
sheet = wb['Sheet3']

row_number = 1  # Change this to the row number you want to print
for col in range(1, sheet.max_column + 1):
    print(sheet.cell(row=row_number, column=col).value, end=" ")
print()
