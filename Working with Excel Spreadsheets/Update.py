# 1. Updating a Specific Cell
# To update the value of a specific cell, you can directly assign a new value to it:

import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('example.xlsx')

# Select the sheet
sheet = wb['Sheet1']  # Change 'Sheet1' to your sheet name

# Update the value in cell A1
sheet['A1'] = 'Updated Value'

# Save the workbook
wb.save('example.xlsx')

# ----------------------------------------------------------------------------------------------------

# 2. Updating Multiple Cells
# You can update multiple cells by assigning new values to each of them:

import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('example.xlsx')

# Select the sheet
sheet = wb['Sheet1']  # Change 'Sheet1' to your sheet name

# Update multiple cells
sheet['A1'] = 'Updated Value 1'
sheet['B1'] = 'Updated Value 2'
sheet['C1'] = 456

# Save the workbook
wb.save('example.xlsx')

# --------------------------------------------------------------------------------------------------

# 3. Updating Cells in a Loop
# If you need to update a range of cells, you can use a loop:

import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('example.xlsx')

# Select the sheet
sheet = wb['Sheet1']  # Change 'Sheet1' to your sheet name

# Update cells in a loop
for i in range(1, sheet.max_row + 1):
    sheet['A' + str(i)] = 'Row ' + str(i)

# Save the workbook
wb.save('example.xlsx')

# --------------------------------------------------------------------------------------------------

# 4. Conditional Updates
# You can also update cells based on certain conditions. For example, updating all cells in a column if they match a specific value:

import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('example.xlsx')

# Select the sheet
sheet = wb['Sheet1']  # Change 'Sheet1' to your sheet name

# Update cells conditionally
for row in range(1, sheet.max_row + 1):
    if sheet['A' + str(row)].value == 'Old Value':
        sheet['A' + str(row)].value = 'New Value'

# Save the workbook
wb.save('example.xlsx')
