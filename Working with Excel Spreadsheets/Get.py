# 1. Getting the Value of a Specific Cell
# To get the value of a specific cell, you can use the cell’s coordinates:

import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('example.xlsx')

# Select the sheet
sheet = wb['Sheet1']  # Change 'Sheet1' to your sheet name

# Get the value of cell A1
value = sheet['A1'].value
print(value)

# -------------------------------------------------------------------------------------------------------------------------

# 2. Getting Values from a Row
# To get all the values from a specific row, you can iterate through the columns in that row:

row_number = 1  # Change this to the row number you want to get
row_values = [sheet.cell(row=row_number, column=col).value for col in range(1, sheet.max_column + 1)]
print(row_values)

# -------------------------------------------------------------------------------------------------------------------------

# 3. Getting Values from a Column
# To get all the values from a specific column, you can iterate through the rows in that column:

column_letter = 'A'  # Change this to the column letter you want to get
column_values = [sheet[column_letter + str(row)].value for row in range(1, sheet.max_row + 1)]
print(column_values)

# -------------------------------------------------------------------------------------------------------------------------

# 4. Getting Values Based on a Condition
# You can also retrieve values based on certain conditions. For example, getting all rows where a specific column matches a given value:

search_value = 'John Doe'  # Change this to the value you are looking for
matching_rows = []

for row in range(1, sheet.max_row + 1):
    if sheet['A' + str(row)].value == search_value:  # Change 'A' to the column you want to search
        row_data = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
        matching_rows.append(row_data)

print(matching_rows)

# -------------------------------------------------------------------------------------------------------------------------

# 5. Getting the Maximum Value in a Column
# To find the maximum value in a specific column, you can iterate through the rows and keep track of the maximum value:

column_letter = 'B'  # Change this to the column letter you want to search
max_value = None

for row in range(1, sheet.max_row + 1):
    cell_value = sheet[column_letter + str(row)].value
    if max_value is None or (cell_value is not None and cell_value > max_value):
        max_value = cell_value

print(max_value)
