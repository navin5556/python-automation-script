import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('example.xlsx')

# Select the sheet
sheet = wb['Sheet1']  # Change 'Sheet1' to your sheet name

# Insert data into specific cells
sheet['A1'] = 'Hello'
sheet['B1'] = 'World'
sheet['C1'] = 123

# Save the workbook
wb.save('example.xlsx')

# -------------------------------------------------------------------------------------------------------------------------

# You can also insert data into cells using row and column indices
# Insert data using row and column indices
sheet.cell(row=2, column=1).value = 'Python'
sheet.cell(row=2, column=2).value = 'Automation'
sheet.cell(row=2, column=3).value = 456

# Save the workbook
wb.save('example.xlsx')

# -------------------------------------------------------------------------------------------------------------------------

# 2. The append method allows you to add a new row of data to the end of the sheet
import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('example.xlsx')

# Select the sheet
sheet = wb['Sheet1']  # Change 'Sheet1' to your sheet name

# Append a new row of data
sheet.append(['New', 'Row', 456])

# Save the workbook
wb.save('example.xlsx')

# -------------------------------------------------------------------------------------------------------------------------

# 3. If you need to insert data at a specific position and shift existing data, you can manually shift the data and 
# then insert the new data. Here’s an example of how to insert a row at a specific position:

import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('example.xlsx')

# Select the sheet
sheet = wb['Sheet1']  # Change 'Sheet1' to your sheet name

# Insert a new row at position 2 (shifting existing rows down)
sheet.insert_rows(2)

# Insert data into the new row
# sheet['A2'] = 'Inserted'
# sheet['B2'] = 'Row'
# sheet['C2'] = 789

# Save the workbook
wb.save('example.xlsx')
