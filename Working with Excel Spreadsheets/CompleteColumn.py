import openpyxl 
wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet3']
      
# column_letter = 'B'  # Change this to the column letter you want to print
# for row in range(1, sheet.max_row + 1):
#     print(sheet[column_letter + str(row)].value)


for row in range(1, sheet.max_row + 1):
    print(sheet.cell(row=row, column=2).value)