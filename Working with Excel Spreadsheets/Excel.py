import openpyxl, os
wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet3']

# print(sheet['A1'].value)

# for cellObj in list(sheet.columns)[1]:
#         print(cellObj.value)

# print(sheet['A'].value)
# print(sheet.cell(row=1, column=2).value)

# for i in range(1,sheet.max_row+1):
#     for j in range(2,sheet.max_column+1):
#         print(sheet.cell(row=i, column=j).value, end=" ")
#     print()
for i in range(1,sheet.max_row+1):
    print(str(sheet['A'+str(i)].value) + " " + str(sheet['B'+str(i)].value) + " " + str(sheet['C'+str(i)].value))
    # print(sheet['B'+str(i)].value, end=" ")
    # print(sheet['c'+str(i)].value, end=" ")
    print()
# print(sheet.max_column)


