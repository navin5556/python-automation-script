# 1. Deleting a Cell’s Value 
# To delete the value of a specific cell, you can simply set it to None:

import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('example.xlsx')

# Select the sheet
sheet = wb['Sheet1']  # Change 'Sheet1' to your sheet name

# Delete the value in cell A1
sheet['A4'].value = None

# Save the workbook
wb.save('example.xlsx')

# -------------------------------------------------------------------------------------------------------------------------


# 2. Deleting an Entire Row
# To delete an entire row, you can use the delete_rows method:

import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('example.xlsx')

# Select the sheet
sheet = wb['Sheet1']  # Change 'Sheet1' to your sheet name

# Delete the second row
sheet.delete_rows(2)

# Save the workbook
wb.save('example.xlsx')

# -------------------------------------------------------------------------------------------------------------------------

# 3. Deleting an Entire Column
# To delete an entire column, you can use the delete_cols method:

import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('example.xlsx')

# Select the sheet
sheet = wb['Sheet1']  # Change 'Sheet1' to your sheet name

# Delete the second column (B)
sheet.delete_cols(2)

# Save the workbook
wb.save('example.xlsx')

# -------------------------------------------------------------------------------------------------------------------------

# 4. Deleting Multiple Rows or Columns
# You can also delete multiple rows or columns by specifying the number of rows or columns to delete:

import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('example.xlsx')

# Select the sheet
sheet = wb['Sheet1']  # Change 'Sheet1' to your sheet name

# Delete 3 rows starting from the second row
sheet.delete_rows(2, 3)

# Delete 2 columns starting from the second column
# sheet.delete_cols(2, 2)

# Save the workbook
wb.save('example.xlsx')
