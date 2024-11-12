# To find the row with the maximum number in a specific column
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet3']

column_letter = 'C'  # Change this to the column letter you want to search
max_value = None
max_row = None

for row in range(1, sheet.max_row + 1):
    cell_value = sheet[column_letter + str(row)].value
    if max_value is None or (cell_value is not None and cell_value > max_value):
        max_value = cell_value
        max_row = row

if max_row is not None:
    for col in range(1, sheet.max_column + 1):
        print(sheet.cell(row=max_row, column=col).value, end=" ")
    print()
else:
    print("No numeric values found in column.")
